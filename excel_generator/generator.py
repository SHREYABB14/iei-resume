import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def clean_special_characters(val):
    if not isinstance(val, str):
        return val
    # Remove common unicode bullets, symbols, and Private Use Area (PUA) characters
    val = re.sub(
        r'[\u2022\u25cf\u25fe\u25c6\u25a0\u25b2\u25ba\u2756\u274b\u2713\u2714\u274c\u2708\u2709\u2700-\u27bf\uf000-\uf8ff\u2605\u2606\u29bf\u2b57\u25cb\u25cc\u25cd\u25ce\u25cf\u25d0-\u25d3\u25e2-\u25e5\u25f8-\u25ff\u2b24\u25a1\u25c7\u25c9\u25ca\u25cb\u25cc]', 
        '', 
        val
    )
    # Strip any remaining leading/trailing bullets, hyphens, stars, spaces, or bullet-like prefixes
    val = re.sub(r'^[\s\-\*•❖☐☑☒▪➢➤➔➜✓+]+', '', val)
    return val.strip()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _thin_border():
    side = Side(style='thin')
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font():
    return Font(name='Arial', bold=True, size=10)


def _cell_font():
    return Font(name='Arial', size=10)


def _header_fill():
    return PatternFill('solid', start_color='BDD7EE')  # light blue


def _write_faculty_meta(ws, faculty_name: str, designation: str, branch: str):
    """
    Write the Name / Designation / Branch header block at the top of a
    publication worksheet (rows 1-3).  Returns the row number where the
    data table should begin.
    """
    labels = [
        ('Name of Faculty', faculty_name),
        ('Designation', designation),
        ('Branch', branch),
    ]
    for row_idx, (label, value) in enumerate(labels, start=1):
        label_cell = ws.cell(row=row_idx, column=1, value=f'{label} :')
        label_cell.font = Font(name='Arial', bold=True, size=10)
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = _cell_font()

    # Return row number after the header block (leave one blank row gap)
    return len(labels) + 2


def _write_pub_table(ws, publications: list, start_row: int):
    """
    Write the publication data table starting at `start_row`.
    Columns: Sr No | Title of Paper | Journal Name |
             Published Under / Journal Details | Year of Publication |
             Impact Factor | Scopus Index
    """
    col_headers = [
        'Sr No',
        'Title of Paper',
        'Journal Name',
        'Published Under / Journal Details',
        'Year of Publication',
        'Impact Factor',
        'Scopus Index',
    ]

    # Write column headers
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _thin_border()

    # Write publication rows
    for sr_no, pub in enumerate(publications, start=1):
        row = start_row + sr_no
        row_data = [
            sr_no,
            pub.get('Title of Paper', ''),
            pub.get('Journal Name', ''),
            pub.get('Published Under / Publisher', ''),
            pub.get('Year of Publication', ''),
            pub.get('Impact Factor', 'N/A'),
            pub.get('Scopus Indexed', 'Unknown'),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = _cell_font()
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = _thin_border()

    # Set reasonable column widths
    col_widths = [8, 45, 30, 35, 20, 15, 15]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[start_row].height = 30


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_excels(master_rows, pub_rows):
    """
    Generate two Excel workbooks:

    1. master_faculty_database.xlsx
       - Single sheet with one row per faculty member.
       - Columns match the NAAC/NBA master data template.

    2. publication_details.xlsx
       - One worksheet per faculty member (sheet name = faculty name).
       - Each sheet contains a 3-row metadata header (Name, Designation,
         Branch) followed by a formatted publication table.
       - Faculty name/designation/branch are NOT repeated per publication row.

    Parameters
    ----------
    master_rows : list[dict]
        Each dict represents one faculty member's master data.
    pub_rows : list[dict]
        Each dict represents one publication record and must include
        'Faculty Name', 'Designation', 'Branch' for sheet grouping,
        plus publication-level fields (title, journal_name, etc.).
    """

    cols_to_strip_numbers = {
        'Name', 'Current Designation', 'Current Department', 'Current Organization',
        'UG Degree', 'UG Branch', 'UG University',
        'PG Degree', 'PG Branch', 'PG University',
        'PhD University', 'Title of Paper', 'Journal Name', 'Published Under / Journal Details',
        'Faculty Name', 'Designation', 'Department', 'Publication Type'
    }

    cleaned_master_rows = []
    for row in master_rows:
        cleaned_row = {}
        for k, v in row.items():
            cleaned_val = clean_special_characters(v)
            if k in cols_to_strip_numbers and isinstance(cleaned_val, str):
                cleaned_val = re.sub(r'^\d+[\.\)\-]?\s+', '', cleaned_val)
            cleaned_row[k] = cleaned_val
        cleaned_master_rows.append(cleaned_row)
    master_rows = cleaned_master_rows

    cleaned_pub_rows = []
    for row in (pub_rows or []):
        cleaned_row = {}
        for k, v in row.items():
            cleaned_val = clean_special_characters(v)
            if k in cols_to_strip_numbers and isinstance(cleaned_val, str):
                cleaned_val = re.sub(r'^\d+[\.\)\-]?\s+', '', cleaned_val)
            cleaned_row[k] = cleaned_val
        cleaned_pub_rows.append(cleaned_row)
    pub_rows = cleaned_pub_rows

    # ==========================================================================
    # 1. MASTER SHEET  (structure unchanged from original)
    # ==========================================================================

    expected_master_cols = [
        'Source File', 'Name', 'Email', 'Phone',
        'Current Designation', 'Current Department', 'Current Organization',
        'UG Degree', 'UG Branch', 'UG University', 'UG Year',
        'PG Degree', 'PG Branch', 'PG University', 'PG Year',
        'PhD University', 'PhD Year',
        'Academic Experience', 'Industry Experience',
        'Research Experience', 'Administrative Experience', 'Total Experience',
        'Journal Count', 'International Conference Count',
        'National Conference Count', 'Book Count', 'Book Chapter Count',
        'Patent Count', 'Projects Count', 'Awards Count', 'Membership Count',
        'FDP Attended', 'STTP Attended',
        'ORCID', 'Google Scholar', 'Confidence Score',
    ]

    master_df = pd.DataFrame(master_rows) if master_rows else pd.DataFrame(columns=expected_master_cols)

    for col in expected_master_cols:
        if col not in master_df.columns:
            master_df[col] = ''

    master_df = master_df[expected_master_cols]

    # ==========================================================================
    # 2. PUBLICATION WORKBOOK (one sheet per faculty)
    # ==========================================================================

    # Group publication records by faculty name so each gets its own worksheet.
    # Preserve insertion order (Python 3.7+ dicts are ordered).
    faculty_pubs: dict[str, list] = {}
    faculty_meta: dict[str, dict] = {}

    for pub in (pub_rows or []):
        name = pub.get('Faculty Name', 'Unknown')
        if name not in faculty_pubs:
            faculty_pubs[name] = []
            faculty_meta[name] = {
                'designation': pub.get('Designation', ''),
                'branch': pub.get('Department', ''),  # main.py sends 'Department', not 'Branch'
            }
        faculty_pubs[name].append(pub)

    pub_wb = Workbook()
    # Remove the default blank sheet openpyxl creates
    default_sheet = pub_wb.active
    if default_sheet is not None:
        pub_wb.remove(default_sheet)

    for faculty_name, pubs in faculty_pubs.items():
        # Sanitise sheet name: Excel limits to 31 chars, no special chars
        sheet_name = faculty_name[:31].replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
        ws = pub_wb.create_sheet(title=sheet_name)

        meta = faculty_meta[faculty_name]
        table_start = _write_faculty_meta(
            ws,
            faculty_name=faculty_name,
            designation=meta['designation'],
            branch=meta['branch'],
        )

        _write_pub_table(ws, pubs, start_row=table_start)

    # If no publications were provided, add a placeholder sheet
    if not faculty_pubs:
        ws = pub_wb.create_sheet(title='No Publications')
        ws['A1'] = 'No publication data available.'

    return master_df, pub_wb